# encoding=utf-8

import glob
from openpyxl import load_workbook
from openpyxl import Workbook

from stat_services import handle_sen, \
        anylizer, \
        get_anylized_word_map, \
        get_sen_list_by_word, \
        get_sen_price_by_sen, \
        get_sen_click_count_pc_by_sen, \
        get_sen_click_count_mobile_by_sen, \
        handle_sen_price_dict, \
        handle_pc_sen_click_count, \
        handle_mobile_sen_click_count, \
        clear_data


WB = Workbook()

exclude_src_file_name = None

def handle(src_file_name, dest_file_name):
    wb = load_workbook(src_file_name)
    source_sheet = None
    for sheet in wb:
        #print (sheet.title)
        source_sheet = sheet.title

    sheet = wb.get_sheet_by_name(source_sheet)
    for row in sheet.rows:
        for cell in row:
            pass
            #print (cell.value)


    for column in sheet.columns:
        for cell in column:
            if cell.value:
                pass

    handle_filter_sheet(sheet, dest_file_name)
    handle_export(sheet, dest_file_name) 


def bulk_insert_col(sheet, row, col, value_list):
    row = row
    for value in value_list:
        sheet.cell(column=col, row=row, value="%s" % (value))
        row = row + 1


def filter_label_list(is_filter=True):
    exclude = get_exclude_list()
    label_list = []
    anylized_workd_map = get_anylized_word_map()
    for word_dict in anylized_workd_map:
        word, info = word_dict
        column_field_name = '-'.join([word, str(info.get('count'))])
        print(word, exclude)
        if is_filter:
            if word not in exclude: 
                label_list.append(column_field_name)
        else:
            label_list.append(column_field_name)
    return label_list


def filter_label_name():
    exclude = get_exclude_list()
    name_list = []
    anylized_workd_map = get_anylized_word_map()
    for word_map in anylized_workd_map:
        word, _= word_map
        if word not in exclude:
            name_list.append(word)
    return name_list
    

def init_data(sheet):
    sen_list = [] 
    for i in range(4, sheet.max_row):
    #for i in range(4, 40):
        if sheet.cell(i, 1).value is not None:
            sen = sheet.cell(i, 1).value.strip() 
            if sen not in sen_list:
                sen_list.append(sen) 
                handle_sen(sen)
                price = sheet.cell(i, 6).value.strip() 
                print('price====', price)
                handle_sen_price_dict(sen, price)
                click_count_pc = sheet.cell(i, 5).value.strip() 
                print('pc_click_count', click_count_pc)
                handle_pc_sen_click_count(sen, click_count_pc)
                click_count_mobile = sheet.cell(i, 4).value.strip() 
                print('mobile_click_count', click_count_mobile)
                handle_mobile_sen_click_count(sen, click_count_mobile)


def add_col(analyze_sheet):
    label_name_list = filter_label_name()
    row = 2
    col = 1
    for word in label_name_list:
        row_data_list = []
        row_data_list = get_sen_list_by_word(word)
        #print('word=%s, row_data_list=%s' % (word, row_data_list))
        if row_data_list:
            bulk_insert_col(analyze_sheet, row, col, row_data_list)
            col = col + 1

        row_sen_price_list = []
        for sen in row_data_list:
            row_sen_price_list.append(get_sen_price_by_sen(sen))
        if row_sen_price_list:
            bulk_insert_col(analyze_sheet, row, col, row_sen_price_list)
            col = col + 1

        row_sen_pc_click_count_list = []
        for sen in row_data_list:
            row_sen_pc_click_count_list.append(get_sen_click_count_pc_by_sen(sen))
        if row_sen_pc_click_count_list:
            bulk_insert_col(analyze_sheet, row, col, row_sen_pc_click_count_list)
            col = col + 1

        row_sen_mobile_click_count_list = []
        for sen in row_data_list:
            row_sen_mobile_click_count_list.append(get_sen_click_count_mobile_by_sen(sen))
        if row_sen_mobile_click_count_list:
            bulk_insert_col(analyze_sheet, row, col, row_sen_mobile_click_count_list)
            col = col + 1


def get_exclude_list():
    global exclude_src_file_name
    exclude_list = []
    try:
        wb = load_workbook(exclude_src_file_name)
        filter_sheet = wb.get_sheet_by_name('filter')
        for row in range(2, filter_sheet.max_row):
            value = filter_sheet.cell(row=row, column=2).value
            if value and value not in exclude_list:
                exclude_list.append(value)
    except Exception:
        pass
    return exclude_list


def handle_export(sheet, dest_file_name):
    wb = load_workbook(dest_file_name)
    analyze_sheet = wb.create_sheet("analyze_word", index=2)

    label_list = filter_label_list()
    
    new_label_list = []
    for label in label_list:
        new_label_list.append(label)
        new_label_list.append('price')
        new_label_list.append('pc_click_count')
        new_label_list.append('mobile_click_count')
    analyze_sheet.append(new_label_list)
    add_col(analyze_sheet)

    analyze_sheet.title = "analyze_word"
    wb.save(dest_file_name)


def handle_filter_sheet(sheet, dest_file_name):
    init_data(sheet)
    try:
        wb = load_workbook(dest_file_name)
        filter_sheet = wb.get_sheet_by_name('filter')
    except Exception as e:
        wb = Workbook()
        filter_sheet = wb.create_sheet('filter', index=1) 

    filter_sheet.cell(row=1, column=1, value='word') 
    filter_sheet.cell(row=1, column=2, value='exclude') 
    word_list = filter_label_list(is_filter=False)
    bulk_insert_col(filter_sheet, 2, 1, word_list)
    wb.save(dest_file_name)

def main():
    global exclude_src_file_name
    path = "*.xlsx"
    for fname in glob.glob(path):
        print(fname)
        if '_anlyzed' not in fname:
            clear_data()
            src_file_name = fname
            dest_file_name = src_file_name.split('.')[0]
            dest_file_name = '%s%s.xlsx' % (dest_file_name, '_anlyzed')
            exclude_src_file_name = dest_file_name
            print('start', src_file_name, dest_file_name)
            handle(src_file_name, dest_file_name)
            print ('success')
        else:
            print ('include self')

if __name__ == '__main__':
    main()

